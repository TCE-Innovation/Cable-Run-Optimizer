from settings import *

if local_code_flag:

    import os
    import re
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    import math
    from cable_classes import *
    import subprocess

    file_path = None


    # Extract the diameter and weight of all cables from Cable Sizes.xlsx
    def get_cable_sizes():
        print("[STATUS] Fetching cable sizes...")

        # Path to the folder containing the Cable Pull Sheet
        file_path = r'C:\Users\roneill\OneDrive - Iovino Enterprises, LLC\Documents 1' \
                    r'\Code\Git Files\Cable-Run-Optimizer\Cable Sizes.xlsx'

        # Load the Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        cable_sizes_list = parse_cable_sizes_excel(sheet)

        # Close the Excel workbook
        workbook.close()

        print(f"[PASS] Cable sizes acquired.\n")

        # for info in cable_sizes_list:
        #     print(f"{info.size}")

        return cable_sizes_list


    # Open cable pull sheet and extract all the cables and their info from it
    def get_cable_pull_sheet(cable_list, cable_sizes_list):

        # Initialize empty cable list to hold cables from pull sheet
        cable_list = []

        print("[STATUS] Fetching cable pull sheet...")
        # Updated column headers to match your fixed column format
        pull_number_col_index = 1  # Column A
        cable_size_col_index = 2  # Column B
        express_col_index = 3  # Column C
        stationing_start_col_index = 4  # Column D
        stationing_end_col_index = 5  # Column E
        distance_col_index = 6  # Column F (for absolute distances)

        # Path to the "Test Basic Pull Sheet.xlsx" file
        file_path = r'C:\Users\roneill\OneDrive - Iovino Enterprises, LLC\Documents 1' \
                    r'\Code\Git Files\Cable-Run-Optimizer\Test Basic Pull Sheet.xlsx'

        # Load the Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Regular expression pattern to detect stationing values (both numerical and location descriptors)
        # stationing_pattern = r'\d+\+\d+|[A-Z\-]+'

        # Iterate over the rows to extract information from relevant columns
        for row in sheet.iter_rows(min_row=2, values_only=True):
            pull_number = row[0]
            cable_size = row[1]
            express = row[2]
            stationing_start = row[3]
            stationing_end = row[4]

            # Check if stationing_start is a numerical stationing value
            if '+' in str(stationing_start):
                # If it's a numerical stationing value, treat it as such
                absolute_distance = None
                # stationing_values_numeric.append((stationing_start, stationing_end))
            else:
                # If it's a location descriptor, check if there's an absolute distance in Column F
                absolute_distance = row[5]  # Assuming Column F contains absolute distance
                stationing_text_pairs.append((stationing_start, stationing_end))  # Log text descriptor pair

            # Find the corresponding CableParameters object based on the cable size
            # Initialize a variable to store cable information
            cable_info = None
            # Iterate through the list of cable size information
            for info in cable_sizes_list:
                # Check if the cable size matches the size of the current cable object
                if info.size == cable_size:
                    # If a match is found, store the cable size information
                    cable_info = info
                    # Exit the loop since we found the relevant cable size
                    break

            if cable_info is not None:
                cable = Cable(
                    pull_number,
                    stationing_start,
                    stationing_end,
                    cable_size,
                    express,
                    cable_info.diameter,
                    cable_info.pounds_per_foot,
                    cable_info.cross_sectional_area,
                    absolute_distance,
                    cable_info.two_conductor,
                    cable_info.length,
                    cable_info.width
                    # cable_info.two_conductor ? True
                )

                cable_list.append(cable)
        print(f"[PASS] Cable pull sheet obtained.\n")
        print(len(cable_list))
        # Print out the cables at the end of the function
        # for cable in cable_list:
        #     print(f"Cable: Pull #{cable.pull_number}, Size: {cable.cable_size}, Express: {cable.express}, "
        #           f"Stationing Start: {cable.stationing_start}, Stationing End: {cable.stationing_end}, "
        #           f"Absolute Distance: {cable.absolute_distance},\n Diameter: {cable.diameter}, "
        #           f"Weight: {cable.weight}, Cross-sectional Area: {cable.cross_sectional_area}, "
        #           f"Two Conductor: {cable.two_conductor}, Length: {cable.length}, Width: {cable.width}, "
        #           f"Radius: {cable.radius}, Angle: {cable.angle}")

        # Return the cable list if needed for further processing
        return cable_list


elif server_code_flag:

    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    import math
    from io import BytesIO
    from .cable_classes import *
    import logging
    from .azure import upload_to_azure


    # Extract the diameter and weight of all cables from Cable Sizes.xlsx
    def get_cable_sizes(cable_sizes, bytes_flag):
        try:
            if bytes_flag:
                logging.info("Bytes flag is true, loading workbook from bytes.")
                workbook = openpyxl.load_workbook(BytesIO(cable_sizes))
            else:
                logging.info("Bytes flag is false, loading workbook from file.")
                workbook = openpyxl.load_workbook(BytesIO(cable_sizes.read()))

            sheet = workbook.active

            logging.info("Running parse_cable_sizes_excel function.")
            cable_sizes_list = parse_cable_sizes_excel(sheet)

            logging.info("Compatible Cables (Cables Sizes.xslx):")
            for cable in cable_sizes_list:
                logging.info("Cable %s", cable.size)

            # Close the workbook
            workbook.close()

            return cable_sizes_list

        except openpyxl.utils.exceptions.InvalidFileException:
            # Handle the case where the file cannot be opened (invalid Excel file)
            logging.info("Invalid Excel file or sheet")
        except Exception as e:
            # Handle other exceptions
            logging.info(f"An error occurred: {str(e)}")


    # Open cable pull sheet and extract all the cables and their info from it
    def get_cable_pull_sheet(pull_sheet, cable_sizes_list):
        cable_list = []

        global stationing_text_pairs
        stationing_text_pairs = []

        # Load the Excel file
        workbook = openpyxl.load_workbook(BytesIO(pull_sheet.read()))
        sheet = workbook.active

        logging.info("Parsing through Cable Pull Sheet...")

        # Iterate over the rows to extract information from relevant columns
        for row in sheet.iter_rows(min_row=2, values_only=True):
            pull_number = row[0]
            cable_size = row[1]
            logging.info("Raw row data (pull sheet): %s", row[1])
            express = row[2]
            stationing_start = row[3]
            stationing_end = row[4]

            # Check if stationing_start is a numerical stationing value
            if '+' in str(stationing_start):
                # If it's a numerical stationing value, treat it as such
                absolute_distance = None
            else:
                # If it's a location descriptor, check if there's an absolute distance in Column F
                absolute_distance = row[5]  # Assuming Column F contains absolute distance
                stationing_text_pairs.append((stationing_start, stationing_end))  # Log text descriptor pair

            # Find the corresponding CableParameters object based on the cable size
            # Initialize a variable to store cable information
            cable_info = None
            # Iterate through the list of cable size information
            for info in cable_sizes_list:
                # Check if the cable size matches the size of the current cable object
                if info.size == cable_size:
                    # If a match is found, store the cable size information
                    cable_info = info
                    logging.info("Match found. %s from CS and %s from PS", info.size, cable_size)
                    # Exit the loop since we found the relevant cable size
                    break

            if cable_info is not None:
                cable = Cable(
                    pull_number,
                    stationing_start,
                    stationing_end,
                    cable_size,
                    express,
                    cable_info.diameter,
                    cable_info.pounds_per_foot,
                    cable_info.cross_sectional_area,
                    absolute_distance,
                    cable_info.two_conductor
                )
                cable_list.append(cable)
                logging.info("Cable %s : %s pulled", cable.pull_number, cable.cable_size)
                logging.info("Length of cable_list: %s", len(cable_list))

        print(f"[PASS] Cable pull sheet obtained.\n")

        # Print out the cables at the end of the function
        for cable in cable_list:
            print(f"Cable: Pull #{cable.pull_number}, Size: {cable.cable_size}, Express: {cable.express}, "
                  f"Stationing Start: {cable.stationing_start}, Stationing End: {cable.stationing_end}, "
                  f"Absolute Distance: {cable.absolute_distance}")

        # Close the workbook
        workbook.close()

        # Return the cable list if needed for further processing
        return cable_list


def sort_stationing(cable_list):
    if server_code_flag:
        logging.info("Running sort_stationing function.")

    global stationing_text_pairs
    stationing_values_numeric = []

    if server_code_flag:
        logging.info("Start: stationing_text_pairs %s", stationing_text_pairs)
        logging.info("Start: stationing_values_numeric %s", stationing_values_numeric)

    # Create a set to store unique stationing values
    unique_stationing_values = set()

    if server_code_flag:
        logging.info("In sort stationing. Length of cable_list: %s", len(cable_list))

    for cable in cable_list:
        if cable.absolute_distance is None:  # Only adding numeric stationing values
            unique_stationing_values.add(cable.stationing_start)

        if cable.absolute_distance is None:  # Only adding numeric stationing values
            unique_stationing_values.add(cable.stationing_end)

    # Convert the set to a list and sort it numerically
    stationing_values_numeric = sorted(list(unique_stationing_values))

    print("[STATUS] Printing all stationing values obtained: ")
    # Print out all the stationing values
    for value in stationing_values_numeric:
        print(value)

    # Create a set to store unique stationing text pairs
    unique_stationing_text_pairs = set(stationing_text_pairs)
    stationing_text_pairs = list(unique_stationing_text_pairs)

    if server_code_flag:
        logging.info("End: stationing_text_pairs %s", stationing_text_pairs)
        logging.info("End: stationing_values_numeric %s", stationing_values_numeric)

    return stationing_values_numeric, stationing_text_pairs


def parse_cable_sizes_excel(sheet):
    # Initialize variables to track the special case
    special_case = False
    first_row_skipped = False  # Flag to skip the first row when in the special case
    length = None
    width = None

    cable_sizes_list = []

    if server_code_flag:
        logging.info("Parsing through Cables Sizes.xlsx...")

    # Iterate over the rows starting from the second row (second because first row has the headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check for the special case
        if row[0] == "* 2 Conductor Cables Below *":
            special_case = True
            continue  # Skip this row

        if special_case:
            if not first_row_skipped:
                first_row_skipped = True
                continue  # Skip the first row

            # This is the case of 2 conductors cables
            size = row[0]
            length = row[1] if row[1] > row[2] else row[2]
            width = row[2] if row[1] > row[2] else row[1]
            weight = row[3]

            # Calculate the cross-sectional area as the product of length and width
            cross_sectional_area = round(length * width, 5)

            # Setting the diameter of the cable to be one conductor + jacket surrounding it
            cable = CableParameters(size, width, weight, cross_sectional_area, True, length, width)
            # print(
            #     f"Created cable size: {cable.size}, Diameter: {cable.diameter}, Weight: {cable.pounds_per_foot}, "
            #     f"Cross-sectional Area: {cable.cross_sectional_area}, \nIs 2C: {cable.two_conductor} "
            #     f"Length: {cable.length}, Width: {cable.width}\n")

            # Create a CableParameters object with diameter set to "None"
            # cable = CableParameters(size, None, weight, cross_sectional_area, True)
            cable_sizes_list.append(cable)

        else:
            # Extract the cable parameters as usual
            size = row[0]
            if server_code_flag:
                logging.info("Raw row data (cable sizes): %s", row[0])
            diameter = row[1]
            pounds_per_foot = row[2]
            cross_sectional_area = round(math.pi * ((float(diameter) / 2) ** 2), 4)

            # Create a CableParameters object and append it to the list
            cable = CableParameters(size, diameter, pounds_per_foot, cross_sectional_area, False)
            # print(f"Created cable size: {cable.size}, Diameter: {cable.diameter}, Weight: {cable.pounds_per_foot}, "
            #       f"Cross-sectional Area: {cable.cross_sectional_area}, Is 2C: {cable.two_conductor}")
            cable_sizes_list.append(cable)

    if server_code_flag:
        logging.info("End of parse_cable_sizes_excel. Cable sizes list length: %s", len(cable_sizes_list))

    return cable_sizes_list


def generate_output_file(cable_run_list, runType):
    if local_code_flag and run_conduit_optimization:
        from cable_classes import conduit_sizes
    elif server_code_flag and run_conduit_optimization:
        from .cable_classes import conduit_sizes

        logging.info("All supported conduit sizes:")
        for i in range(9):
            logging.info(conduit_sizes[i])

    if server_code_flag:
        logging.info("Generating output file...")
    elif local_code_flag:
        print("\n[STATUS] Generating output file...")

    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    if runType == "Conduit":
        # Set column headers
        headers = [
            "Conduit",
            "Stationing Start",
            "Stationing End",
            "Pull #",
            "Cable Size",
            "Express/Local",
            "Minimum Conduit Size (in)",
            "Conduit Fill",
            "Upsized Conduit (in)",
            "Upsized Conduit Fill"
        ]

        sheet.append(headers)

        # Write conduit data and cable attributes to the Excel file
        for conduit_name, conduit in cable_run_list.items():

            # Get corresponding index in conduit_sizes for current conduit size
            # This is to be able to show upsized conduit
            # conduit_sizes_index = next((int(index) for index, x in enumerate(conduit_sizes) if x == conduit.conduit_size),
            #                            None)
            for i in range(len(conduit_sizes) - 1):
                if conduit_sizes[i] == conduit.conduit_size:
                    conduit_sizes_index = i
                    break

            for cable in conduit.cables:
                row_data = [
                    f"Conduit {conduit.conduit_number}",  # Conduit name
                    f"{conduit.stationing_start}",  # Stationing start
                    f"{conduit.stationing_end}",  # Stationing end
                    int(cable.pull_number),  # Pull Number
                    cable.cable_size,  # Cable size (ex. 7C#14)
                    cable.express,  # Express or local
                    conduit.conduit_size,  # Conduit size in inches
                    str(conduit.conduit_fill) + "%",  # f"{round((100 - conduit_free_air_space), 2)}%",   # Conduit fill
                    conduit_sizes[conduit_sizes_index + 1],  # Upsized conduit
                    f"{(100 * conduit.conduit_area / (math.pi * ((conduit_sizes[conduit_sizes_index + 1] / 2) ** 2))):.2f}%"
                    # Upsized conduit fill
                ]
                sheet.append(row_data)
    elif runType == "Messenger":
        # Set column headers
        headers = [
            "Bundle",
            "Stationing Start",
            "Stationing End",
            "Pull #",
            "Cable Size",
            "Express/Local",
            "Bundle Diameter (in)",
            "Bundle Weight (lb/mft)",
        ]

        sheet.append(headers)

        # Write bundle data and cable attributes to the Excel file
        for bundle_name, bundle in cable_run_list.items():

            for cable in bundle.cables:
                row_data = [
                    f"Bundle {bundle.bundle_number}",  # Bundle number
                    f"{bundle.stationing_start}",  # Stationing start
                    f"{bundle.stationing_end}",  # Stationing end
                    int(cable.pull_number),  # Individual cable pull number
                    cable.cable_size,  # Cable size (ex. 7C#14)
                    cable.express,  # Express or local
                    round(bundle.bundle_diameter, 2),  # Bundle diameter
                    round(bundle.bundle_weight / 1000, 2)  # Bundle weight
                ]
                sheet.append(row_data)

    if server_code_flag:
        logging.info("Conduit/bundle list length: %s", len(cable_run_list))
    elif local_code_flag:
        print(f"Conduit/bundle list length: {len(cable_run_list)}")

    # Set column width to fit the text in each header
    for col_num, header in enumerate(headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        column_width = max(len(header), max(len(str(cell.value)) for cell in sheet[col_letter]))
        sheet.column_dimensions[col_letter].width = column_width + 2  # Adding some extra width for padding

    # Apply bold font to the header row
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Dictionary to store counts for each conduit/bundle name
    run_counts = {}

    # Count the occurrences of each conduit/bundle name
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)):
        run_name = row[0].value
        if run_name not in run_counts:
            run_counts[run_name] = 1
        else:
            run_counts[run_name] += 1

    # Merge cells for each conduit/bundle name and adjust Stationing Start and Stationing End columns
    for run_name, count in run_counts.items():
        start_row = None
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1), start=2):
            if row[0].value == run_name:
                if start_row is None:
                    start_row = row_num
                if count == 1:
                    continue

                if row_num - start_row + 1 == count:

                    # Define the column letters for merging and alignment
                    columns_to_merge = ['A', 'B', 'C', 'F', 'G', 'H', 'I', 'J']

                    # Loop through the columns and apply merging and alignment
                    for column in columns_to_merge:
                        column_range = f'{column}{start_row}:{column}{row_num}'
                        sheet.merge_cells(column_range)
                        sheet[f'{column}{start_row}'].alignment = Alignment(vertical='center', horizontal='center')

                    start_row = None

    # Center align cells in all columns for rows starting from the second row
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center')

    if local_code_flag:
        # Save the workbook to a file
        output_filename = "Output File.xlsx"
        workbook.save(output_filename)

        print(f"[PASS] Output file {output_filename} has been saved.")
        print(f"[STATUS] Opening output file...")

        pdf_file_path = 'C:/Users/roneill/OneDrive - Iovino Enterprises, LLC/Documents 1/Code/Git Files/Cable-Run-Optimizer/Output File.xlsx'

        # subprocess.run(["start", "", pdf_file_path], shell=True, check=True)

    elif server_code_flag:

        # Upload the workbook to azure blob storage
        sas_url = upload_to_azure(workbook)

        print(f"Conduit/bundle data has been saved to uploaded to blob storage.")

        return sas_url
