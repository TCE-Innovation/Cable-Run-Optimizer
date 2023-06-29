import os
import openpyxl
from tabulate import tabulate

def process_excel_files(folder_path):
    # Iterate over files in the folder
    for file_name in os.listdir(folder_path):
        # Check if the file is an Excel file
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            # Construct the full file path
            file_path = os.path.join(folder_path, file_name)

            # Load the Excel file
            workbook = openpyxl.load_workbook(file_path)

            # Print the sheet names
            print(f"File: {file_name}")
            print("Sheet Names:")
            for sheet_name in workbook.sheetnames:
                print(sheet_name)

            # Iterate over sheets and print cell values
            for sheet in workbook:
                print(f"\nSheet: {sheet.title}")
                data = []
                for row in sheet.iter_rows(values_only=True):
                    filtered_row = [cell_value if cell_value is not None else "" for cell_value in row]
                    data.append(filtered_row)
                headers = data[0]
                data = data[1:]
                print(tabulate(data, headers=headers, tablefmt="grid"))

            # Close the workbook
            workbook.close()


# Provide the path to the folder containing the Excel files
folder_path = r'C:\Users\roneill\Documents\CRO'

# Call the function to process Excel files in the folder
process_excel_files(folder_path)
